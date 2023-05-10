import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
prod_list = inv_file["Sheet1"]

prod_per_supp = {}
total_val_per_supp = {}     # added while finding the inventory details
prod_under_10_inv = {}      # added to find under 10 inventory

# print(prod_list.max_row)

for prod_row in range(2, prod_list.max_row + 1):
    supp_name = prod_list.cell(prod_row, 4).value
    inv = prod_list.cell(prod_row, 2).value
    price = prod_list.cell(prod_row, 3).value
    prod_num = prod_list.cell(prod_row, 1).value
    inv_price = prod_list.cell(prod_row, 5)

    # calculating no of products per supplier
    if supp_name in prod_per_supp:
        current_no_prod = prod_per_supp.get(supp_name)
        prod_per_supp[supp_name] = current_no_prod + 1
    else:
        prod_per_supp[supp_name] = 1

    # calculating total value of inventory per supplier
    if supp_name in total_val_per_supp:
        current_total_value = total_val_per_supp.get(supp_name)
        total_val_per_supp[supp_name] = current_total_value + (inv * price)
    else:
        total_val_per_supp[supp_name] = inv * price

    # calculating products which are less than 10 in the inventory
    if inv < 10:
        prod_under_10_inv[int(prod_num)] = int(inv)

    # adding a new column for total inventory price
    inv_price.value = inv * price


print(prod_per_supp)
print(total_val_per_supp)
print(prod_under_10_inv)
inv_file.save("inventory_with_total_value.xlsx")